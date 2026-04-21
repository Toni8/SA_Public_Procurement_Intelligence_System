"""
============================================================
 reports/excel_reporter.py
 Automated Excel Intelligence Report Generator

 Produces a fully formatted .xlsx workbook with 6 sheets:
   1. Executive Summary   — KPIs + policy brief
   2. Opportunity Matrix  — ranked province×quarter×sector
   3. Heatmap             — province × quarter pivot (colour-coded)
   4. Anomaly Register    — all flagged contracts
   5. Value Benchmarks    — median/mean by sector+province
   6. Model Performance   — ML comparison metrics

 Run: python reports/excel_reporter.py
============================================================
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import logging

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import DATA_DIR, MODEL_DIR, REPORT_DIR

logger = logging.getLogger(__name__)

# ── Colour palette ──────────────────────────────────────────
NAVY     = "1E3A5F"
GOLD     = "C9A227"
LIGHT_BG = "F0F4F8"
GREEN    = "1A7A4A"
RED      = "C0392B"
ORANGE   = "E67E22"
WHITE    = "FFFFFF"

HEADER_FONT  = Font(name="Arial", bold=True, color=WHITE, size=11)
HEADER_FILL  = PatternFill("solid", start_color=NAVY)
SUBHDR_FILL  = PatternFill("solid", start_color="2E5F8A")
SUBHDR_FONT  = Font(name="Arial", bold=True, color=WHITE, size=10)
ALT_FILL     = PatternFill("solid", start_color=LIGHT_BG)
BODY_FONT    = Font(name="Arial", size=10)
KPI_FILL     = PatternFill("solid", start_color="E8F4FD")
TITLE_FONT   = Font(name="Arial", bold=True, color=NAVY, size=14)
SECTION_FONT = Font(name="Arial", bold=True, color=NAVY, size=11)

THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
)

ZAR_FMT   = 'R #,##0'
ZAR_M_FMT = 'R #,##0.0,,"M"'
PCT_FMT   = '0.00%'
INT_FMT   = '#,##0'


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _header_row(ws, row: int, values: list, fill=None, font=None):
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _data_row(ws, row: int, values: list, alt: bool = False):
    fill = ALT_FILL if alt else PatternFill("solid", start_color=WHITE)
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


# ─────────────────────────────────────────────────────────────
# Sheet 1: Executive Summary
# ─────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, df: pd.DataFrame, brief: dict):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50
    ws.row_dimensions[1].height = 40

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "🏛  SOUTH AFRICA — PUBLIC PROCUREMENT INTELLIGENCE REPORT"
    t.font  = TITLE_FONT
    t.fill  = PatternFill("solid", start_color="E8F4FD")
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:C2")
    sub = ws["A2"]
    sub.value = "Source: eTenders SA (OCDS Format) | Median-anchored reporting"
    sub.font  = Font(name="Arial", italic=True, color="666666", size=10)
    sub.alignment = Alignment(horizontal="center")

    # KPI section
    ws.merge_cells("A4:C4")
    kpi_hdr = ws["A4"]
    kpi_hdr.value = "KEY PERFORMANCE INDICATORS"
    kpi_hdr.font  = SECTION_FONT
    kpi_hdr.fill  = SUBHDR_FILL
    kpi_hdr.font  = SUBHDR_FONT
    kpi_hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    es = brief.get("executive_summary", {})
    kpis = [
        ("Total Contracts Analysed",  es.get("total_contracts", ""),        "Number of OCDS records with positive contract value"),
        ("Total Procurement Spend",   es.get("total_spend", ""),            "Cumulative contract value (ZAR)"),
        ("Median Contract Value",     es.get("median_contract", ""),        "Preferred central tendency — mean is misleading (skew ratio shown below)"),
        ("Mean Contract Value",       es.get("mean_contract", ""),          ""),
        ("Mean / Median Ratio",       es.get("mean_median_ratio", ""),      "Extreme right skew — EDA finding: 314× gap in real data"),
        ("Open Tender Rate",          es.get("open_tender_pct", ""),        "% of tenders issued via competitive open bidding"),
        ("Top Province by Spend",     es.get("top_province", ""),           ""),
        ("Top Sector by Spend",       es.get("top_sector", ""),             ""),
    ]

    _header_row(ws, 5, ["Indicator", "Value", "Notes"])
    for r, (label, value, note) in enumerate(kpis, 6):
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=1).fill = KPI_FILL
        ws.cell(row=r, column=2, value=value).font = BODY_FONT
        ws.cell(row=r, column=2).fill = KPI_FILL
        ws.cell(row=r, column=3, value=note).font = Font(name="Arial", italic=True, size=9, color="555555")
        ws.cell(row=r, column=3).fill = KPI_FILL
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = THIN_BORDER

    # Key findings
    row = len(kpis) + 8
    ws.merge_cells(f"A{row}:C{row}")
    hdr = ws[f"A{row}"]
    hdr.value = "KEY FINDINGS FROM EDA"
    hdr.font  = SUBHDR_FONT
    hdr.fill  = SUBHDR_FILL
    hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    findings = brief.get("key_findings", [])
    for i, f in enumerate(findings, row + 1):
        ws.merge_cells(f"A{i}:C{i}")
        cell = ws[f"A{i}"]
        cell.value = f"• {f}"
        cell.font  = BODY_FONT
        cell.fill  = ALT_FILL if i % 2 == 0 else PatternFill("solid", start_color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        cell.border = THIN_BORDER
        ws.row_dimensions[i].height = 30

    # Recommendations
    row2 = row + len(findings) + 2
    ws.merge_cells(f"A{row2}:C{row2}")
    hdr2 = ws[f"A{row2}"]
    hdr2.value = "STRATEGIC RECOMMENDATIONS"
    hdr2.font  = SUBHDR_FONT
    hdr2.fill  = PatternFill("solid", start_color=GREEN)
    hdr2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    recs = brief.get("recommendations", [])
    for i, rec in enumerate(recs, row2 + 1):
        ws.merge_cells(f"A{i}:C{i}")
        cell = ws[f"A{i}"]
        cell.value = f"✓  {rec}"
        cell.font  = Font(name="Arial", size=10, color=GREEN)
        cell.fill  = PatternFill("solid", start_color="E8F5E9")
        cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        cell.border = THIN_BORDER
        ws.row_dimensions[i].height = 28


# ─────────────────────────────────────────────────────────────
# Sheet 2: Opportunity Matrix
# ─────────────────────────────────────────────────────────────

def _build_opportunity_sheet(ws, matrix: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    headers = [
        "Rank", "Province", "Sector", "Quarter",
        "Contracts", "Median Value (R)", "Total Market (R)",
        "Reliability", "Opportunity Score", "% National Spend", "Gauteng Flag"
    ]
    _set_col_widths(ws, {
        "A": 7, "B": 20, "C": 12, "D": 20,
        "E": 11, "F": 18, "G": 18, "H": 18,
        "I": 18, "J": 17, "K": 14
    })

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "PROVINCE × SECTOR × QUARTER — OPPORTUNITY MATRIX"
    t.font  = TITLE_FONT
    t.fill  = PatternFill("solid", start_color="E8F4FD")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = "Sorted by Opportunity Score = log(median_value) × reliability_weight × seasonal_weight"
    sub.font  = Font(name="Arial", italic=True, size=9, color="666666")
    sub.alignment = Alignment(horizontal="center")

    _header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 30

    reliability_colours = {
        "High Confidence": "1A7A4A",
        "Reliable":        "2980B9",
        "Caution":         "E67E22",
        "Insufficient Data": "C0392B",
    }

    for r, (_, row) in enumerate(matrix.iterrows(), 4):
        alt  = r % 2 == 0
        vals = [
            r - 3,
            row.get("province", ""),
            row.get("category", ""),
            row.get("quarter_label", ""),
            row.get("contract_count", 0),
            row.get("median_value", 0),
            row.get("total_value", 0),
            row.get("reliability_tier", ""),
            row.get("opportunity_score", 0),
            row.get("pct_of_national_spend", 0) / 100,
            "⚠ Yes" if row.get("gauteng_flag", 0) else "No",
        ]
        _data_row(ws, r, vals, alt=alt)

        # Format number columns
        ws.cell(r, 5).number_format  = INT_FMT
        ws.cell(r, 6).number_format  = ZAR_FMT
        ws.cell(r, 7).number_format  = ZAR_FMT
        ws.cell(r, 9).number_format  = "0.000"
        ws.cell(r, 10).number_format = PCT_FMT

        # Colour-code reliability
        rel = str(row.get("reliability_tier", ""))
        colour = reliability_colours.get(rel, "AAAAAA")
        ws.cell(r, 8).font = Font(name="Arial", bold=True, color=colour, size=10)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{3 + len(matrix)}"


# ─────────────────────────────────────────────────────────────
# Sheet 3: Heatmap
# ─────────────────────────────────────────────────────────────

def _build_heatmap_sheet(ws, matrix: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    from models.opportunity_matrix import build_heatmap_pivot
    pivot = build_heatmap_pivot(matrix, value_col="median_value")
    pivot = pivot / 1_000_000   # Convert to R millions

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "PROVINCE × QUARTER HEATMAP — Median Contract Value (R Millions)"
    t.font  = TITLE_FONT
    t.fill  = PatternFill("solid", start_color="E8F4FD")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.cell(3, 1, "Province")
    ws.cell(3, 1).font = HEADER_FONT
    ws.cell(3, 1).fill = HEADER_FILL
    ws.cell(3, 1).alignment = Alignment(horizontal="center")

    for col_i, col_name in enumerate(pivot.columns, 2):
        ws.cell(3, col_i, col_name)
        ws.cell(3, col_i).font  = HEADER_FONT
        ws.cell(3, col_i).fill  = HEADER_FILL
        ws.cell(3, col_i).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_i)].width = 18

    ws.column_dimensions["A"].width = 22

    for row_i, (prov, row_data) in enumerate(pivot.iterrows(), 4):
        ws.cell(row_i, 1, prov)
        ws.cell(row_i, 1).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row_i, 1).fill = ALT_FILL if row_i % 2 == 0 else PatternFill("solid", start_color=WHITE)
        ws.cell(row_i, 1).border = THIN_BORDER

        for col_i, val in enumerate(row_data, 2):
            cell = ws.cell(row_i, col_i, round(float(val), 2))
            cell.number_format = '#,##0.0'
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Colour scale conditional formatting
    start_row, end_row = 4, 4 + len(pivot) - 1
    end_col = get_column_letter(1 + len(pivot.columns))
    color_range = f"B4:{end_col}{end_row}"

    ws.conditional_formatting.add(
        color_range,
        ColorScaleRule(
            start_type="min",  start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="FFF3CD",
            end_type="max",    end_color="C0392B",
        )
    )

    ws.merge_cells(f"A{end_row + 2}:F{end_row + 2}")
    note = ws[f"A{end_row + 2}"]
    note.value = "⚠ Gauteng values are suppressed in opportunity scoring due to volume/value anomaly (EDA finding)."
    note.font  = Font(name="Arial", italic=True, color="E67E22", size=9)


# ─────────────────────────────────────────────────────────────
# Sheet 4: Anomaly Register
# ─────────────────────────────────────────────────────────────

def _build_anomaly_sheet(ws, df: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    flagged = df[df.get("anomaly_flag", pd.Series(["Normal"] * len(df))) != "Normal"].copy()

    headers = ["OCID", "Province", "Sector", "Method", "Supplier",
               "Contract Value (R)", "Tender Date", "Anomaly Flag", "Severity"]
    _set_col_widths(ws, {
        "A": 18, "B": 18, "C": 12, "D": 12, "E": 25,
        "F": 18, "G": 14, "H": 45, "I": 10
    })

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = f"PROCUREMENT ANOMALY REGISTER — {len(flagged):,} Flagged Contracts"
    t.font  = Font(name="Arial", bold=True, color=WHITE, size=13)
    t.fill  = PatternFill("solid", start_color=RED)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    _header_row(ws, 2, headers)

    sev_colours = {
        5: RED, 4: RED, 3: ORANGE, 2: "E6A817", 1: "2980B9", 0: GREEN
    }

    flagged_sorted = flagged.sort_values("anomaly_severity", ascending=False) \
        if "anomaly_severity" in flagged.columns else flagged

    for r, (_, row) in enumerate(flagged_sorted.iterrows(), 3):
        alt  = r % 2 == 0
        sev  = int(row.get("anomaly_severity", 0))
        vals = [
            str(row.get("ocid", "")),
            row.get("province", ""),
            row.get("category", ""),
            row.get("method", ""),
            row.get("supplier_name", ""),
            row.get("contract_value", 0),
            str(row.get("tender_date", ""))[:10],
            row.get("anomaly_flag", ""),
            sev,
        ]
        _data_row(ws, r, vals, alt=alt)
        ws.cell(r, 6).number_format = ZAR_FMT
        ws.cell(r, 8).font = Font(name="Arial", bold=True,
                                  color=sev_colours.get(sev, "000000"), size=10)

    ws.freeze_panes = "A3"
    if len(flagged) > 0:
        ws.auto_filter.ref = f"A2:I{2 + len(flagged)}"


# ─────────────────────────────────────────────────────────────
# Sheet 5: Value Benchmarks
# ─────────────────────────────────────────────────────────────

def _build_benchmarks_sheet(ws, df: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    grp = (
        df[df["contract_value"] > 0]
        .groupby(["category", "province"])
        .agg(
            sample_size  = ("contract_value", "count"),
            median_value = ("contract_value", "median"),
            mean_value   = ("contract_value", "mean"),
            min_value    = ("contract_value", "min"),
            max_value    = ("contract_value", "max"),
            std_value    = ("contract_value", "std"),
        )
        .reset_index()
    )
    grp = grp[grp["sample_size"] >= 5].sort_values("median_value", ascending=False)

    headers = ["Sector", "Province", "Sample Size", "Median (R)",
               "Mean (R)", "Min (R)", "Max (R)", "Std Dev (R)", "Reliability"]
    _set_col_widths(ws, {
        "A": 12, "B": 20, "C": 13, "D": 16,
        "E": 16, "F": 16, "G": 16, "H": 16, "I": 14
    })

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "CONTRACT VALUE BENCHMARKS BY SECTOR × PROVINCE"
    t.font  = TITLE_FONT
    t.fill  = PatternFill("solid", start_color="E8F4FD")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    note = ws["A2"]
    note.value = "All estimates anchored to MEDIAN — mean is misleading due to extreme right skew (EDA finding: 314× mean/median gap)"
    note.font  = Font(name="Arial", italic=True, size=9, color="C0392B")
    note.alignment = Alignment(horizontal="center")

    _header_row(ws, 3, headers)

    for r, (_, row) in enumerate(grp.iterrows(), 4):
        n = int(row["sample_size"])
        reliability = "High" if n >= 10 else "Medium" if n >= 5 else "Low"
        vals = [
            row["category"], row["province"], n,
            row["median_value"], row["mean_value"],
            row["min_value"], row["max_value"],
            row.get("std_value", 0), reliability
        ]
        _data_row(ws, r, vals, alt=r % 2 == 0)
        for col in [4, 5, 6, 7, 8]:
            ws.cell(r, col).number_format = ZAR_FMT

        rel_colour = {"High": GREEN, "Medium": "2980B9", "Low": ORANGE}[reliability]
        ws.cell(r, 9).font = Font(name="Arial", bold=True, color=rel_colour, size=10)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{3 + len(grp)}"


# ─────────────────────────────────────────────────────────────
# Sheet 6: Model Performance
# ─────────────────────────────────────────────────────────────

def _build_model_sheet(ws):
    ws.sheet_view.showGridLines = False

    model_csv = MODEL_DIR / "model_comparison.csv"
    if model_csv.exists():
        model_df = pd.read_csv(model_csv)
    else:
        model_df = pd.DataFrame([
            {"Model": "Models not trained yet — run train_pipeline.py first"}
        ])

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "ML MODEL PERFORMANCE COMPARISON"
    t.font  = TITLE_FONT
    t.fill  = PatternFill("solid", start_color="E8F4FD")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    note = ws["A2"]
    note.value = "Target: log1p(contract_value) | Higher R² = better fit | Lower MAPE % = better ZAR accuracy"
    note.font  = Font(name="Arial", italic=True, size=9, color="666666")
    note.alignment = Alignment(horizontal="center")

    if len(model_df.columns) > 1:
        _header_row(ws, 3, list(model_df.columns))
        for r, (_, row) in enumerate(model_df.iterrows(), 4):
            _data_row(ws, r, list(row.values), alt=r % 2 == 0)
            for col_i, val in enumerate(row.values, 1):
                if isinstance(val, float):
                    ws.cell(r, col_i).number_format = "0.0000"

        _set_col_widths(ws, {get_column_letter(i+1): 16 for i in range(len(model_df.columns))})
    else:
        ws.cell(4, 1, model_df.iloc[0, 0])
        ws.cell(4, 1).font = Font(name="Arial", italic=True, color=ORANGE)

    # EDA methodology notes
    start = 3 + len(model_df) + 3
    ws.merge_cells(f"A{start}:F{start}")
    ws[f"A{start}"].value = "MODELLING METHODOLOGY NOTES"
    ws[f"A{start}"].font  = SECTION_FONT
    ws[f"A{start}"].fill  = SUBHDR_FILL
    ws[f"A{start}"].font  = SUBHDR_FONT

    notes = [
        "Log-transformation of target (log1p) applied throughout — EDA confirmed extreme right skew.",
        "Interaction features: province × quarter × sector — EDA confirmed no single strong linear predictor.",
        "Ensemble = VotingRegressor of XGBoost + LightGBM + RandomForest.",
        "All value estimates should be reported as median-anchored with reliability tier noted.",
        "Gauteng is modelled separately due to anomalously low value-per-tender profile.",
    ]
    for i, note in enumerate(notes, start + 1):
        ws.merge_cells(f"A{i}:F{i}")
        cell = ws[f"A{i}"]
        cell.value = f"• {note}"
        cell.font  = BODY_FONT
        cell.fill  = ALT_FILL if i % 2 == 0 else PatternFill("solid", start_color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        cell.border = THIN_BORDER
        ws.row_dimensions[i].height = 26


# ─────────────────────────────────────────────────────────────
# Master export function
# ─────────────────────────────────────────────────────────────

def generate_excel_report(
    df: pd.DataFrame,
    matrix: pd.DataFrame,
    brief: dict,
    output_path: Path = None,
) -> Path:
    """
    Build the full Excel intelligence report.
    Returns the path to the saved .xlsx file.
    """
    if output_path is None:
        output_path = REPORT_DIR / "Procurement_Intelligence_Report.xlsx"

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet 1 — Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    _build_summary_sheet(ws1, df, brief)
    logger.info("  ✓ Sheet 1: Executive Summary")

    # Sheet 2 — Opportunity Matrix
    ws2 = wb.create_sheet("Opportunity Matrix")
    _build_opportunity_sheet(ws2, matrix)
    logger.info("  ✓ Sheet 2: Opportunity Matrix")

    # Sheet 3 — Heatmap
    ws3 = wb.create_sheet("Province Heatmap")
    _build_heatmap_sheet(ws3, matrix)
    logger.info("  ✓ Sheet 3: Province Heatmap")

    # Sheet 4 — Anomaly Register
    ws4 = wb.create_sheet("Anomaly Register")
    _build_anomaly_sheet(ws4, df)
    logger.info("  ✓ Sheet 4: Anomaly Register")

    # Sheet 5 — Benchmarks
    ws5 = wb.create_sheet("Value Benchmarks")
    _build_benchmarks_sheet(ws5, df)
    logger.info("  ✓ Sheet 5: Value Benchmarks")

    # Sheet 6 — Model Performance
    ws6 = wb.create_sheet("Model Performance")
    _build_model_sheet(ws6)
    logger.info("  ✓ Sheet 6: Model Performance")

    wb.save(output_path)
    logger.info(f"Excel report saved → {output_path}")
    return output_path


# ─────────────────────────────────────────────────────────────
# CLI runner
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import json
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s — %(message)s")

    logger.info("Loading data for Excel report...")
    csv = DATA_DIR / "sample.csv"
    if not csv.exists():
        from utils.generate_sample_data import generate
        df = generate(n=5000)
    else:
        df = pd.read_csv(csv, parse_dates=["tender_date"])

    from models.anomaly_detector import run_full_anomaly_pipeline, anomaly_summary_report
    df = run_full_anomaly_pipeline(df)
    summary = anomaly_summary_report(df)

    from models.opportunity_matrix import build_opportunity_matrix
    matrix = build_opportunity_matrix(df)

    from models.recommendation_engine import generate_policy_brief
    brief = generate_policy_brief(df, matrix, summary)

    out = generate_excel_report(df, matrix, brief)
    print(f"\n✅ Excel report: {out}")

